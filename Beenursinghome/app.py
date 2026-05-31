import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, time

st.set_page_config(page_title="요양원 수급자 상담일지 자동생성기", layout="wide")

st.title("🏥 요양원 수급자 상담일지 자동생성기 v3")
st.caption("상담영역 → 문제상황 → 원인 → 조치 → 급여반영 → 엑셀 다운로드")

REQUIRED_ITEMS = [
    "수급자 성명", "생년월일", "성별", "장기요양 등급", "입소일", "생활실 호수", "본인부담률",
    "상담 일시", "상담 구분", "상담 장소", "상담 방법", "상담 대상자 성명", "상담자와의 관계",
    "기관 담당자 직책 및 성명", "동석자", "상담 목적 및 사유",
    "신체 기능 상태 변화 및 건강 관련 요구", "인지 기능 및 심리 상태", "식사 및 영양 상태",
    "배설 상태 및 개인 위생", "일상생활 및 여가 활동 요구", "가족 및 보호자의 요청 사항",
    "기관의 설명 및 안내 사항", "문제 해결 또는 제공된 정보 내용", "향후 급여 제공 계획",
    "급여 반영 여부", "사후 조치 및 모니터링 계획", "보호자 전달/공지사항",
    "상담자 서명", "대상자 또는 보호자 서명"
]

AREA_DATA = {
    "식사상담": {
        "problems": ["식사량 감소", "폭식", "편식", "식사 거부", "연하곤란", "저작곤란", "식사 집중력 저하", "잔반 증가", "식사 중 사레"],
        "causes": ["구강 불편", "치아 문제", "소화 불편", "기분 저하", "음식 선호도 변화", "질병 영향", "인지 저하", "피로감"],
        "actions": ["죽 제공", "진밥 제공", "잡곡밥 조정", "경관식 확인", "식사 보조 강화", "선호 음식 반영", "간호사 공유", "보호자 안내"],
    },
    "배변상담": {
        "problems": ["변비", "설사", "배변 횟수 감소", "배변 실수", "복부 불편감", "화장실 이동 어려움"],
        "causes": ["수분섭취 부족", "활동량 감소", "식사량 감소", "약물 영향", "인지 저하", "배변 습관 변화"],
        "actions": ["수분섭취 권장", "배변 관찰기록", "간호사 공유", "식사량 확인", "활동량 증가 유도", "보호자 안내"],
    },
    "수면상담": {
        "problems": ["야간 각성", "불면", "낮잠 증가", "수면 중 불안", "새벽 배회", "낮 시간 피로"],
        "causes": ["환경 변화", "불안감", "통증", "배뇨 욕구", "치매 증상", "낮 활동 부족"],
        "actions": ["수면환경 조정", "낮 활동량 증가", "야간 관찰 강화", "간호사 공유", "보호자 안내", "안정적 일과 제공"],
    },
    "기저귀케어 상담": {
        "problems": ["기저귀 불편감", "피부 발적", "습진 우려", "교체 거부", "배뇨량 변화", "배변 후 피부자극"],
        "causes": ["장시간 착용", "피부 민감", "배뇨·배변 증가", "인지 저하", "개인위생 거부"],
        "actions": ["교체 주기 확인", "피부상태 관찰", "청결관리 강화", "피부보호제 검토", "간호사 공유", "보호자 안내"],
    },
    "낙상예방 상담": {
        "problems": ["보행 불안정", "지팡이 보행", "워커 사용", "휠체어 이동", "침상 이동 위험", "화장실 이동 위험"],
        "causes": ["근력저하", "어지러움", "인지 저하", "야간 이동", "환경 장애물", "보조기구 사용 미숙"],
        "actions": ["이동 보조 강화", "침상 주변 정리", "화장실 동행", "낙상위험 교육", "보조기구 점검", "직원 공유"],
    },
    "욕창예방 상담": {
        "problems": ["와상 상태", "피부 발적", "체위변경 어려움", "엉덩이 압박", "피부 건조", "통증 호소"],
        "causes": ["장시간 동일 자세", "영양 저하", "기저귀 착용", "활동량 부족", "피부 약화"],
        "actions": ["체위변경 강화", "피부관찰", "영양상태 확인", "기저귀케어 연계", "간호사 보고", "욕창예방 관리"],
    },
    "정서상담": {
        "problems": ["우울감", "불안감", "외로움", "무기력", "짜증 증가", "눈물 보임", "대화 회피"],
        "causes": ["보호자 부재", "환경 변화", "건강 저하", "프로그램 흥미 저하", "대인관계 어려움"],
        "actions": ["정서적 지지", "대화시간 확대", "보호자와 공유", "선호활동 연결", "관찰기록 유지", "프로그램 참여 유도"],
    },
    "프로그램 상담": {
        "problems": ["참여 저조", "흥미 부족", "집중력 저하", "활동 거부", "신체활동 어려움", "인지활동 어려움"],
        "causes": ["건강상태 저하", "인지 저하", "선호도 불일치", "피로감", "대인관계 부담"],
        "actions": ["선호 프로그램 반영", "참여 방식 조정", "짧은 활동 제공", "개별 격려", "참여도 기록", "급여계획 반영 검토"],
    },
    "치매·인지상태 상담": {
        "problems": ["기억력 저하", "반복 질문", "시간·장소 혼동", "배회", "망상 의심", "불안·초조", "공격적 언행", "개인위생 거부"],
        "causes": ["치매 증상 진행", "환경 변화", "수면 부족", "신체 불편감", "의사소통 어려움", "일과 변화"],
        "actions": ["짧고 반복적인 설명", "안정적인 환경 유지", "자극 줄이기", "배회 관찰 강화", "보호자 공유", "상태변화기록 연계"],
    },
}

def make_sentence(area, problems, causes, actions, age, grade, mobility, cognition):
    problem_text = ", ".join(problems)
    cause_text = ", ".join(causes)
    action_text = ", ".join(actions)

    consult = (
        f"수급자는 만 {age}세, 장기요양 {grade}이며 현재 이동상태는 '{mobility}', 인지상태는 '{cognition}'으로 확인된다.\n\n"
        f"{area} 관련 상담을 실시한 결과, 주요 문제상황으로 {problem_text}이/가 확인되었다. "
        f"해당 문제는 수급자의 현재 신체기능, 인지상태, 일상생활 적응상태와 연관될 수 있어 세부 확인이 필요하였다.\n\n"
        f"상담 과정에서 {cause_text} 등이 원인으로 작용할 가능성을 확인하였다. "
        f"기관에서는 수급자의 상태가 단순 일시적 변화인지, 반복적으로 나타나는 생활상의 어려움인지 구분하기 위해 관찰을 지속하기로 하였다."
    )

    action = (
        f"{area} 상담 결과에 따라 {action_text} 등의 조치를 실시하거나 검토하기로 하였다. "
        f"해당 내용은 관련 직원과 공유하여 급여제공 과정에서 참고하도록 하며, 필요 시 보호자에게 안내하기로 하였다.\n\n"
        f"향후 동일 문제가 반복되거나 상태 변화가 심화될 경우 상태변화기록지와 연계하여 기록하고, 급여제공계획 반영 여부를 검토하기로 하였다."
    )

    reflection = (
        "상담을 통해 확인된 수급자 및 보호자의 요구사항은 급여제공 과정에 반영 여부를 검토한다. "
        "반영 가능한 사항은 식사, 프로그램, 일상생활 지원, 건강관리, 안전관리 등에 적용하고, "
        "미반영 사항은 사유를 기록하여 추후 상담 시 재확인하기로 한다."
    )

    monitoring = (
        "사후 모니터링은 담당 직원이 일상 관찰을 통해 진행하며, 필요 시 간호 담당자, 요양보호사, 사회복지사가 함께 공유한다. "
        "다음 상담 시 문제상황의 변화 여부와 조치 효과를 재확인하기로 한다."
    )

    return consult, action, reflection, monitoring

def make_excel(data, consult, action, reflection, monitoring, checked_items, missing_items):
    wb = Workbook()
    ws = wb.active
    ws.title = "수급자 상담일지"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    current_row = 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws.cell(current_row, 1).value = f"( {data['기관명']} ) 수급자 상담일지"
    ws.cell(current_row, 1).font = Font(bold=True, size=22)
    ws.cell(current_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 36
    current_row += 1

    info_rows = [
        ("수급자 성명", data["수급자 성명"], "생년월일", data["생년월일"], "성별", data["성별"]),
        ("나이", data["나이"], "장기요양 등급", data["장기요양 등급"], "입소일", str(data["입소일"])),
        ("생활실", data["생활실"], "본인부담률", data["본인부담률"], "상담일자", str(data["상담일자"])),
        ("상담시간", str(data["상담시간"]), "상담구분", data["상담구분"], "상담장소", data["상담장소"]),
        ("상담방법", data["상담방법"], "상담대상자", data["상담대상자"], "관계", data["관계"]),
        ("담당자", data["담당자"], "직책", data["직책"], "동석자", data["동석자"]),
        ("상담영역", data["상담영역"], "이동상태", data["이동상태"], "인지상태", data["인지상태"]),
    ]

    for row_data in info_rows:
        ws.cell(current_row, 1).value = row_data[0]
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).value = row_data[1]
        ws.cell(current_row, 4).value = row_data[2]
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        ws.cell(current_row, 5).value = row_data[3]
        ws.cell(current_row, 7).value = row_data[4]
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=9)
        ws.cell(current_row, 8).value = row_data[5]

        for col in [1, 4, 7]:
            ws.cell(current_row, col).fill = header_fill
            ws.cell(current_row, col).font = Font(bold=True)
            ws.cell(current_row, col).alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    sections = [
        ("상담 내용", consult, 8),
        ("조치 내용", action, 7),
        ("급여제공 반영 여부", reflection, 4),
        ("사후 조치 및 모니터링 계획", monitoring, 4),
    ]

    for title, text, height in sections:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(current_row, 1).value = title
        ws.cell(current_row, 1).fill = header_fill
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.cell(current_row, 1).alignment = Alignment(horizontal="center")
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + height - 1, end_column=9)
        ws.cell(current_row, 1).value = text
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        current_row += height


    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=9)

    ws.cell(current_row, 1).value = "상담자 서명"
    ws.cell(current_row, 4).value = f"{data['담당자']}  (서명)"
    ws.cell(current_row, 7).value = "대상자/보호자  (서명)"
    ws.cell(current_row, 1).fill = header_fill
    ws.cell(current_row, 7).fill = header_fill

    max_row = current_row

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "left",
                vertical=cell.alignment.vertical or "center",
                wrap_text=True
            )

    for r in range(1, max_row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 24
    # 기본 인쇄 설정
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_options.horizontalCentered = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

col1, col2, col3 = st.columns(3)

with col1:
    facility = st.text_input("기관명", "BeeCare 요양원")
    name = st.text_input("수급자 성명", "홍길동")
    birth = st.text_input("생년월일", "1940-01-01")
    gender = st.selectbox("성별", ["남", "여"])
    age = st.number_input("나이", min_value=60, max_value=110, value=85)

with col2:
    grade = st.selectbox("장기요양 등급", ["1등급", "2등급", "3등급", "4등급", "5등급", "인지지원등급"])
    admission_date = st.date_input("입소일", date.today())
    room = st.text_input("생활실 호수", "101호")
    copay = st.selectbox("본인부담률", ["0%", "6%", "9%", "12%", "15%", "20%"])
    consult_date = st.date_input("상담일자", date.today())

with col3:
    consult_time = st.time_input("상담시간", time(10, 0))
    consult_type = st.selectbox("상담 구분", ["정기 상담", "수시 상담", "신규 입소 4주 집중 상담", "보호자 요청 상담"])
    place = st.selectbox("상담 장소", ["상담실", "생활실", "프로그램실", "유선상담", "기타"])
    method = st.selectbox("상담 방법", ["대면 면담", "유선/전화 상담", "온라인/SNS", "기타"])
    target = st.text_input("상담 대상자 성명", "보호자")

col4, col5, col6 = st.columns(3)

with col4:
    relation = st.selectbox("상담자와의 관계", ["본인", "아들", "딸", "배우자", "며느리", "사위", "기타"])
    staff_position = st.selectbox("기관 담당자 직책", ["사회복지사", "간호사", "간호조무사", "요양보호사", "시설장"])
    staff_name = st.text_input("기관 담당자 성명", "사회복지사")

with col5:
    companion = st.text_input("동석자", "없음")
    mobility = st.selectbox("이동상태", ["독립보행", "지팡이 보행", "워커 보행", "휠체어 이동", "부축 필요", "와상 상태"])
    cognition = st.selectbox("인지상태", ["인지 양호", "경도 치매", "중등도 치매", "중증 치매", "의사소통 어려움"])

with col6:
    area = st.selectbox("상담영역 선택", list(AREA_DATA.keys()))

st.subheader("문제상황 / 원인 / 조치 선택")

problems = st.multiselect("문제상황 선택", AREA_DATA[area]["problems"], default=AREA_DATA[area]["problems"][:2])
causes = st.multiselect("상담하게 된 원인 선택", AREA_DATA[area]["causes"], default=AREA_DATA[area]["causes"][:2])
actions = st.multiselect("조치내용 선택", AREA_DATA[area]["actions"], default=AREA_DATA[area]["actions"][:2])

consult, action, reflection, monitoring = make_sentence(area, problems, causes, actions, age, grade, mobility, cognition)

st.subheader("상담 내용")
consult = st.text_area("상담내용 수정 가능", consult, height=220)

st.subheader("조치 내용")
action = st.text_area("조치내용 수정 가능", action, height=180)

st.subheader("급여제공 반영 여부")
reflection = st.text_area("급여제공 반영 내용 수정 가능", reflection, height=130)

st.subheader("사후 조치 및 모니터링 계획")
monitoring = st.text_area("모니터링 계획 수정 가능", monitoring, height=130)

checked_items = []
st.subheader("필수 기재항목 체크리스트")

for item in REQUIRED_ITEMS:
    if st.checkbox(item, value=True, key=f"check_{item}"):
        checked_items.append(item)

missing_items = [item for item in REQUIRED_ITEMS if item not in checked_items]
score = int((len(checked_items) / len(REQUIRED_ITEMS)) * 100)

if score == 100:
    st.success(f"상담일지 완성도: {score}점 / 누락 항목 없음")
elif score >= 85:
    st.warning(f"상담일지 완성도: {score}점 / 일부 항목 확인 필요")
else:
    st.error(f"상담일지 완성도: {score}점 / 필수항목 보완 필요")

data = {
    "기관명": facility,
    "수급자 성명": name,
    "생년월일": birth,
    "성별": gender,
    "나이": age,
    "장기요양 등급": grade,
    "입소일": admission_date,
    "생활실": room,
    "본인부담률": copay,
    "상담일자": consult_date,
    "상담시간": consult_time,
    "상담구분": consult_type,
    "상담장소": place,
    "상담방법": method,
    "상담대상자": target,
    "관계": relation,
    "담당자": staff_name,
    "직책": staff_position,
    "동석자": companion,
    "상담영역": area,
    "이동상태": mobility,
    "인지상태": cognition,
}

excel_file = make_excel(data, consult, action, reflection, monitoring, checked_items, missing_items)

st.download_button(
    label="📥 수급자 상담일지 엑셀 다운로드",
    data=excel_file,
    file_name=f"{name}_{area}_수급자상담일지.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
