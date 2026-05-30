import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.pagebreak import Break
from io import BytesIO
from datetime import date, time
import tempfile

st.set_page_config(page_title="교육일지 생성기", layout="centered")

st.title("교육일지 생성기")

center_name = st.text_input("센터 이름", placeholder="예: 행복주간보호센터")
education_date = st.date_input("교육 날짜", value=date.today())
start_time = st.time_input("시작 시간", value=time(13, 0))
end_time = st.time_input("종료 시간", value=time(14, 0))
education_place = st.text_input("교육장소", placeholder="예: 프로그램실")
teacher_name = st.text_input("교육강사", placeholder="예: 시설장 / 담당자")

doc_type = st.selectbox(
    "일지 종류 선택",
    [
        "노인인권보호교육 일지",
        "노인학대예방교육 일지",
        "소방안전교육 일지",
        "재난대응훈련 일지",
    ]
)

education_style = st.radio(
    "교육내용 유형 선택",
    ["기본교육형", "사례중심형", "실무중심형"],
    horizontal=True
)

content_data = {
    "노인인권보호교육 일지": {
        "기본교육형": {
            "content": """1. 노인의 존엄성과 인권의 기본 개념
2. 어르신의 자기결정권과 선택권 존중
3. 개인정보 및 사생활 보호
4. 언어적·정서적 학대 예방
5. 인권침해 발생 시 보고 절차""",
            "practice": """1. 돌봄 제공 전 어르신의 의사를 먼저 확인한다.
2. 반말, 명령조, 큰소리 사용을 피하고 존댓말을 사용한다.
3. 목욕, 배변, 옷 갈아입기 시 사생활을 보호한다.
4. 프로그램 참여와 식사 선택을 강요하지 않는다.
5. 인권침해 의심 상황은 즉시 담당자에게 보고한다.""",
            "evaluation": "참석자들이 노인인권보호의 기본 원칙을 이해하였으며, 어르신의 자기결정권과 사생활 보호를 현장에서 실천하기로 하였다."
        },
        "사례중심형": {
            "content": """1. 돌봄 현장에서 발생할 수 있는 인권침해 사례
2. 반말, 재촉, 무시 등 언어적 침해 사례
3. 사생활 보호가 필요한 목욕·배변·환복 상황
4. 어르신 선택권이 제한되는 사례
5. 인권침해 의심 시 보고와 기록 방법""",
            "practice": """1. 반복 질문에도 차분히 응대한다.
2. 어르신의 거부 의사를 무시하지 않는다.
3. 신체 노출 상황에서는 가림막과 문 닫기를 실천한다.
4. 개인 물품과 개인정보를 함부로 다루지 않는다.
5. 사례 발생 시 사실 중심으로 기록하고 보고한다.""",
            "evaluation": "참석자들이 실제 사례를 통해 인권침해 가능성을 점검하였고, 일상 돌봄 과정에서 존중하는 태도와 신속한 보고가 필요함을 확인하였다."
        },
        "실무중심형": {
            "content": """1. 종사자의 인권보호 실천수칙
2. 돌봄 전 설명과 동의 절차
3. 개인정보 보호 및 기록 관리
4. 인권침해 의심 상황 관찰
5. 내부 보고 및 재발방지 절차""",
            "practice": """1. 모든 돌봄은 설명 후 동의를 구하고 진행한다.
2. 상담 내용과 개인정보는 외부에 노출하지 않는다.
3. 이상 징후 발견 시 관리자에게 즉시 알린다.
4. 어르신의 선택과 생활습관을 최대한 존중한다.
5. 종사자 간 공유를 통해 재발 방지 방법을 마련한다.""",
            "evaluation": "참석자들이 현장 실무에서 지켜야 할 인권보호 절차를 확인하였으며, 지속적인 관찰과 기록, 보고체계를 강화하기로 하였다."
        },
    },
    "노인학대예방교육 일지": {
        "기본교육형": {
            "content": """1. 노인학대의 정의와 주요 유형
2. 신체적·정서적·언어적 학대 이해
3. 방임 및 경제적 학대 예방
4. 학대 의심 징후 관찰
5. 신고의무와 기관 내부 보고 절차""",
            "practice": """1. 어르신의 상처, 표정, 행동 변화를 관찰한다.
2. 강압적인 말투나 위협적인 행동을 하지 않는다.
3. 식사, 위생, 투약 등 기본 돌봄을 소홀히 하지 않는다.
4. 금전이나 물품을 임의로 사용하지 않는다.
5. 학대 의심 상황은 혼자 판단하지 않고 즉시 보고한다.""",
            "evaluation": "참석자들이 노인학대의 유형과 신고 절차를 이해하였으며, 학대 예방을 위해 세심한 관찰과 신속한 보고가 필요함을 확인하였다."
        },
        "사례중심형": {
            "content": """1. 돌봄 현장에서 발생할 수 있는 노인학대 사례
2. 정서적 학대와 언어적 학대 예방
3. 방임으로 이어질 수 있는 돌봄 누락 사례
4. 경제적 학대 예방과 금전·물품 관리
5. 학대 의심 상황 발견 시 대응 방법""",
            "practice": """1. 짜증, 비난, 협박 표현을 사용하지 않는다.
2. 어르신을 무시하거나 소외시키지 않는다.
3. 식사, 투약, 위생 상태를 정기적으로 확인한다.
4. 어르신의 금전과 물품은 기관 절차에 따라 관리한다.
5. 의심 상황은 사실 중심으로 기록하고 즉시 보고한다.""",
            "evaluation": "참석자들이 사례를 통해 노인학대가 일상 돌봄 중에도 발생할 수 있음을 이해하였으며, 예방을 위한 관찰과 기록의 중요성을 공유하였다."
        },
        "실무중심형": {
            "content": """1. 종사자의 노인학대 예방 역할
2. 학대 의심 징후 체크 방법
3. 기본 돌봄 누락 예방
4. 내부 보고 및 신고 절차
5. 재발 방지와 후속 점검""",
            "practice": """1. 담당 어르신의 신체·정서 상태를 매일 관찰한다.
2. 돌봄 누락이 발생하지 않도록 동료와 확인한다.
3. 학대 의심 상황은 관리자에게 즉시 보고한다.
4. 기록은 사실과 추측을 구분하여 작성한다.
5. 보고 후 후속 조치와 재발 방지 내용을 확인한다.""",
            "evaluation": "참석자들이 노인학대 예방을 위한 실무 절차를 확인하였으며, 돌봄 누락 방지와 신속한 보고체계를 실천하기로 하였다."
        },
    },
    "소방안전교육 일지": {
        "기본교육형": {
            "content": """1. 화재 발생 원인과 예방 수칙
2. 전열기구 및 콘센트 안전관리
3. 소화기 위치와 사용 방법
4. 화재 발생 시 신고 및 초기 대응
5. 비상구와 대피로 확인""",
            "practice": """1. 전열기구 사용 후 전원을 끈다.
2. 콘센트 과부하를 예방한다.
3. 소화기 위치와 압력 상태를 확인한다.
4. 화재 발견 시 큰소리로 알리고 119에 신고한다.
5. 비상구 주변에 물건을 쌓아두지 않는다.""",
            "evaluation": "참석자들이 화재 예방 수칙과 초기 대응 방법을 이해하였으며, 정기적인 소방안전 점검의 필요성을 확인하였다."
        },
        "사례중심형": {
            "content": """1. 전열기구 사용 부주의 사례
2. 콘센트 과부하로 인한 화재 위험 사례
3. 비상구 적치물로 인한 대피 지연 사례
4. 거동불편 어르신 대피 지원 사례
5. 화재 발생 후 인원 확인 방법""",
            "practice": """1. 전기제품 주변을 수시로 점검한다.
2. 위험요소 발견 시 즉시 관리자에게 알린다.
3. 대피로와 비상구를 항상 확보한다.
4. 거동불편 어르신을 우선 대피시킨다.
5. 대피 후 인원과 건강 상태를 확인한다.""",
            "evaluation": "참석자들이 화재 사례를 통해 예방점검과 대피로 확보의 중요성을 이해하였으며, 어르신 대피 지원 절차를 확인하였다."
        },
        "실무중심형": {
            "content": """1. 소방안전 일상 점검 방법
2. 소화기 사용 절차
3. 화재 발생 시 역할 분담
4. 어르신 대피 유도 방법
5. 대피 후 보고와 기록""",
            "practice": """1. 소화기와 비상구 위치를 정기적으로 확인한다.
2. 화재 발생 시 알림, 신고, 대피 순서로 대응한다.
3. 담당 구역별로 어르신 대피를 지원한다.
4. 초기 진압이 어려우면 즉시 대피한다.
5. 대피 후 인원 확인 결과를 기록한다.""",
            "evaluation": "참석자들이 화재 발생 시 역할과 대응 순서를 숙지하였으며, 실제 상황에서 침착하게 대피를 지원하기로 하였다."
        },
    },
    "재난대응훈련 일지": {
        "기본교육형": {
            "content": """1. 지진, 태풍, 폭염, 감염병 등 재난 유형
2. 재난 발생 시 기본 행동요령
3. 어르신 안전 확인 방법
4. 비상연락망 활용
5. 훈련 후 평가와 개선""",
            "practice": """1. 재난 발생 시 침착하게 상황을 알린다.
2. 담당 어르신의 안전을 먼저 확인한다.
3. 필요한 경우 안전한 장소로 이동한다.
4. 비상연락망을 활용해 보호자와 관계기관에 연락한다.
5. 훈련 후 미흡한 점을 기록하고 개선한다.""",
            "evaluation": "참석자들이 재난 유형별 기본 대응 방법을 이해하였으며, 어르신 안전 확보와 비상연락체계의 중요성을 확인하였다."
        },
        "사례중심형": {
            "content": """1. 지진 발생 시 대피 사례
2. 폭염 시 온열질환 예방 사례
3. 감염병 의심 증상 발생 사례
4. 태풍 등 기상재난 대비 사례
5. 비상연락망 활용 사례""",
            "practice": """1. 지진 시 머리를 보호하고 안전한 곳으로 이동한다.
2. 폭염 시 실내 온도와 수분 섭취를 확인한다.
3. 발열이나 호흡기 증상이 있으면 즉시 보고한다.
4. 기상특보 시 외부활동을 제한한다.
5. 보호자 연락과 기관 보고를 신속히 실시한다.""",
            "evaluation": "참석자들이 재난 사례를 통해 상황별 대응 방법을 확인하였으며, 어르신 상태 관찰과 신속한 연락체계가 중요함을 이해하였다."
        },
        "실무중심형": {
            "content": """1. 재난 발생 시 종사자 역할 분담
2. 어르신 이동 및 대피 지원
3. 비상연락망 점검
4. 재난 후 건강상태 확인
5. 훈련 결과 기록과 보완""",
            "practice": """1. 담당 구역과 담당 어르신을 확인한다.
2. 이동이 어려운 어르신은 보조기구를 활용해 지원한다.
3. 비상연락망을 최신 상태로 유지한다.
4. 대피 후 인원과 건강상태를 확인한다.
5. 훈련 결과와 개선사항을 기록한다.""",
            "evaluation": "참석자들이 재난 대응 실무 절차를 확인하였으며, 역할 분담과 반복 훈련을 통해 대응력을 높이기로 하였다."
        },
    },
}

selected = content_data[doc_type][education_style]

st.info(f"선택된 구성: {doc_type} / {education_style}")

st.subheader("참석자 명단 입력")

facility_manager = st.text_input("시설장")
social_worker = st.text_input("사회복지사")
care_worker = st.text_input("요양보호사")
nurse = st.text_input("간호조무사")
physical_therapist = st.text_input("물리치료사")
recipient_1 = st.text_input("수급자 1")
recipient_2 = st.text_input("수급자 2")
recipient_3 = st.text_input("수급자 3")

st.subheader("교육 내용 확인 및 수정")

education_topic = st.text_input("교육 주제", value=doc_type.replace(" 일지", ""))

main_content = st.text_area("주요교육내용", value=selected["content"], height=140)
practice_examples = st.text_area("현장 실천 방법", value=selected["practice"], height=140)
evaluation = st.text_area("교육 결과평가", value=selected["evaluation"], height=90)

uploaded_images = st.file_uploader(
    "교육사진 업로드 (1~2장)",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True
)

if uploaded_images and len(uploaded_images) > 2:
    st.warning("교육사진은 최대 2장까지만 업로드하세요. 앞의 2장만 엑셀에 들어갑니다.")

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = doc_type[:25]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    gray_fill = PatternFill("solid", fgColor="EDEDED")

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    center_title = center_name if center_name else "                    "

    ws.merge_cells("A1:F4")
    ws["A1"] = f"( {center_title} ) {doc_type}"
    ws["A1"].font = Font(size=26, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G1:H1")
    ws["G1"] = "담당자"
    ws["G1"].font = Font(bold=True)
    ws["G1"].fill = gray_fill
    ws["G1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("I1:J1")
    ws["I1"] = "시설장"
    ws["I1"].font = Font(bold=True)
    ws["I1"].fill = gray_fill
    ws["I1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G2:H4")
    ws.merge_cells("I2:J4")

    ws.merge_cells("A5:B5")
    ws["A5"] = "교육일자"
    ws.merge_cells("C5:D5")
    ws["C5"] = str(education_date)

    ws.merge_cells("E5:F5")
    ws["E5"] = "교육시간"
    ws.merge_cells("G5:H5")
    ws["G5"] = f"{start_time.strftime('%H:%M')} ~ {end_time.strftime('%H:%M')}"
    ws.merge_cells("I5:J5")

    ws.merge_cells("A6:B6")
    ws["A6"] = "교육장소"
    ws.merge_cells("C6:D6")
    ws["C6"] = education_place

    ws.merge_cells("E6:F6")
    ws["E6"] = "교육강사"
    ws.merge_cells("G6:J6")
    ws["G6"] = teacher_name

    ws.merge_cells("A7:B7")
    ws["A7"] = "교육내용"
    ws.merge_cells("C7:J7")
    ws["C7"] = education_topic

    ws.merge_cells("A8:J8")

    ws.merge_cells("A9:J9")
    ws["A9"] = "주요교육내용"
    ws["A9"].font = Font(bold=True)
    ws["A9"].fill = gray_fill
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A10:J18")
    ws["A10"] = main_content
    ws["A10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A19:J19")

    ws.merge_cells("A20:J20")
    ws["A20"] = "현장 실천 방법"
    ws["A20"].font = Font(bold=True)
    ws["A20"].fill = gray_fill
    ws["A20"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A21:J27")
    ws["A21"] = practice_examples
    ws["A21"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A28:J28")
    ws["A28"] = "교육 결과평가"
    ws["A28"].font = Font(bold=True)
    ws["A28"].fill = gray_fill
    ws["A28"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A29:J31")
    ws["A29"] = evaluation
    ws["A29"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    photo_start_row = 33
    image_files = uploaded_images[:2] if uploaded_images else []

    if image_files:
        ws.merge_cells(start_row=photo_start_row, start_column=1, end_row=photo_start_row, end_column=10)
        ws.cell(photo_start_row, 1).value = "교육사진"
        ws.cell(photo_start_row, 1).font = Font(bold=True)
        ws.cell(photo_start_row, 1).fill = gray_fill
        ws.cell(photo_start_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        for idx, uploaded_file in enumerate(image_files):
            temp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            temp.write(uploaded_file.getvalue())
            temp.close()

            img = XLImage(temp.name)
            img.width = 250
            img.height = 160

            if idx == 0:
                ws.add_image(img, f"A{photo_start_row + 2}")
            else:
                ws.add_image(img, f"F{photo_start_row + 2}")

        attendee_start_row = photo_start_row + 11
    else:
        attendee_start_row = 32


    ws.merge_cells(start_row=attendee_start_row, start_column=1, end_row=attendee_start_row, end_column=10)
    ws.cell(attendee_start_row, 1).value = "참석자 명단"
    ws.cell(attendee_start_row, 1).font = Font(bold=True)
    ws.cell(attendee_start_row, 1).fill = gray_fill
    ws.cell(attendee_start_row, 1).alignment = Alignment(horizontal="center", vertical="center")

    participants = [
        ("시설장", facility_manager),
        ("사회복지사", social_worker),
        ("요양보호사", care_worker),
        ("간호조무사", nurse),
        ("물리치료사", physical_therapist),
        ("수급자", "    ".join([x for x in [recipient_1, recipient_2, recipient_3] if x])),
    ]

    row = attendee_start_row + 1

    for role, name in participants:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)

        ws.cell(row, 1).value = role
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = gray_fill
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row, 3).value = name
        ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")

        row += 1

    max_row = row - 1

    # 전체 테두리만 적용
    for row_cells in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=10):
        for cell in row_cells:
            cell.border = border

    # 라벨 정렬
    label_cells = [
        "A5", "E5", "A6", "E6", "A7",
        "G1", "I1", "A9", "A20", "A28"
    ]

    for cell_ref in label_cells:
        ws[cell_ref].font = Font(bold=True)
        ws[cell_ref].fill = gray_fill
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 입력값 정렬
    value_cells = ["C5", "G5", "C6", "G6", "C7"]
    for cell_ref in value_cells:
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 본문은 반드시 왼쪽 정렬
    ws["A10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A21"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A29"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    widths = [10, 10, 12, 12, 10, 10, 12, 12, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 23

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[10].height = 150
    ws.row_dimensions[21].height = 130
    ws.row_dimensions[29].height = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if st.button("엑셀 다운로드 파일 만들기"):
    excel_file = create_excel()
    st.download_button(
        label="엑셀 다운로드",
        data=excel_file,
        file_name=f"{doc_type}_{education_style}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
