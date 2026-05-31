import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, time

st.set_page_config(page_title="주간보호센터 상담일지 자동생성기", layout="wide")

st.title("🌿 주간보호센터 상담일지 자동생성기")
st.caption("상담영역 → 문제상황 → 원인 → 조치 → 후속문서 안내 → 엑셀 다운로드")

REQUIRED_ITEMS = [
    "수급자 성명", "생년월일", "성별", "장기요양 등급", "이용시작일", "이용요일", "본인부담률",
    "상담 일시", "상담 구분", "상담 장소", "상담 방법", "상담 대상자 성명", "상담자와의 관계",
    "기관 담당자 직책 및 성명", "동석자", "상담 목적 및 사유",
    "신체 기능 및 건강상태", "인지 기능 및 심리상태", "식사 및 영양상태",
    "배설상태 및 개인위생", "일상생활 및 프로그램 참여", "가족 및 보호자 요청사항",
    "기관 안내사항", "문제 해결 또는 제공 정보", "향후 급여 제공 계획",
    "급여 반영 여부", "사후 조치 및 모니터링 계획", "보호자 전달사항",
    "상담자 서명", "대상자 또는 보호자 서명"
]

AREA_DATA = {
    "식사상담": {
        "problems": ["식사량 감소", "편식", "식사 거부", "식사 속도 저하", "연하곤란", "저작곤란", "잔반 증가", "식사 중 사레"],
        "causes": ["구강 불편", "치아 문제", "소화 불편", "기분 저하", "음식 선호도 변화", "인지 저하", "피로감"],
        "actions": ["죽 제공", "진밥 제공", "잡곡밥 조정", "식사 보조 강화", "선호 음식 반영", "간호사 공유", "보호자 안내"],
    },
    "배변상담": {
        "problems": ["변비", "설사", "배변 횟수 감소", "배변 실수", "복부 불편감", "화장실 이용 어려움"],
        "causes": ["수분섭취 부족", "활동량 감소", "식사량 감소", "약물 영향", "인지 저하", "배변 습관 변화"],
        "actions": ["수분섭취 권장", "배변 관찰기록", "식사량 확인", "활동량 증가 유도", "간호사 공유", "보호자 안내"],
    },
    "송영상담": {
        "problems": ["등원 거부", "하원 불안", "차량 탑승 어려움", "보행 불안정", "보호자 인계 지연", "송영 중 안전 우려"],
        "causes": ["인지 저하", "불안감", "이동 피로", "보행기능 저하", "가족 일정 변화"],
        "actions": ["송영시간 조율", "차량 탑승 보조", "보호자 인계 확인", "직원 공유", "안전관리 강화"],
    },
    "프로그램 참여 상담": {
        "problems": ["참여 저조", "흥미 부족", "집중력 저하", "활동 거부", "신체활동 어려움", "인지활동 어려움"],
        "causes": ["건강상태 저하", "인지 저하", "선호도 불일치", "피로감", "대인관계 부담"],
        "actions": ["선호 프로그램 반영", "참여 방식 조정", "짧은 활동 제공", "개별 격려", "참여도 기록", "급여계획 반영 검토"],
    },
    "행사 참여 상담": {
        "problems": ["생일잔치 참여", "어버이날 행사 참여", "크리스마스 행사 참여", "명절행사 참여", "나들이 행사 참여", "기타 행사 참여"],
        "causes": ["사회적 교류 필요", "정서지원 필요", "가족관계 강화", "생활 활력 제공", "기관 행사 참여 욕구"],
        "actions": ["행사 참여 지원", "활동 반응 기록", "보호자에게 행사 내용 공유", "사진 및 참여도 확인", "향후 행사 참여 독려"],
    },
    "병원동행 상담": {
        "problems": ["내과 진료", "정형외과 진료", "신경과 진료", "치과 진료", "안과 진료", "한의원 진료", "건강검진", "응급진료"],
        "causes": ["정기진료 필요", "통증 호소", "복약 확인 필요", "보호자 요청", "건강상태 변화", "진료일정 도래"],
        "actions": ["병원동행 지원", "진료결과 보호자 공유", "복약사항 확인", "간호사 공유", "다음 진료일정 확인"],
    },
    "개인활동지원 상담": {
        "problems": ["주민센터 방문", "은행 업무", "우체국 업무", "관공서 방문", "장기요양 관련 업무", "기타 외출지원"],
        "causes": ["행정업무 필요", "보호자 요청", "본인 요청", "서류 처리 필요", "개인 일정 발생"],
        "actions": ["외출지원 일정 조율", "이동 안전 확인", "필요서류 안내", "보호자와 일정 공유", "업무처리 결과 기록"],
    },
    "급여계약·인정서 상담": {
        "problems": ["급여계약 연장", "인정서 갱신", "등급 변경", "서비스 시간 변경", "계약 종료 예정", "급여제공계획 변경 필요"],
        "causes": ["계약기간 만료", "인정유효기간 만료", "등급변경 통보", "보호자 요청", "서비스 이용시간 조정"],
        "actions": ["계약 연장 조치", "인정서 갱신 안내", "급여제공계획 유지", "급여제공계획 변경 검토", "욕구사정 실시 안내", "보호자 서명 확인"],
    },
    "병원입원·전원·퇴소 상담": {
        "problems": ["병원 입원 상담", "장기 입원 예정", "퇴원 후 재이용 상담", "타 기관 전원 상담", "서비스 일시중단", "퇴소 상담"],
        "causes": ["건강상태 악화", "입원 치료 필요", "장기요양 이용 중단 필요", "보호자 요청", "돌봄환경 변화"],
        "actions": ["서비스 일시중단 안내", "퇴원 후 재이용 상담", "퇴소 절차 안내", "계약 종료 확인", "관련 기록 정리", "보호자 연락 유지"],
    },
    "방문요양 연계 상담": {
        "problems": ["방문요양 병행 이용", "주간보호 결석일 방문요양 필요", "가족 돌봄 공백", "병원진료 후 가정 내 돌봄 필요", "서비스 시간 조정 필요", "방문요양 전환 검토"],
        "causes": ["가정 내 돌봄 공백", "보호자 근무 일정", "건강상태 변화", "주간보호 이용일 조정", "가족 요청"],
        "actions": ["방문요양 이용 가능 여부 안내", "급여한도 확인", "보호자와 일정 조율", "급여제공계획 변경 검토", "욕구사정 재확인 안내"],
    },
}
def make_followup_notice(area, problems, actions):
    trigger_words = [
        "급여계약 연장",
        "인정서 갱신",
        "등급 변경",
        "서비스 시간 변경",
        "급여제공계획 변경 필요",
        "계약 종료 예정",
        "방문요양 병행 이용",
        "방문요양 전환 검토",
        "서비스 일시중단",
        "퇴소 상담",
        "타 기관 전원 상담",
    ]

    selected_text = " ".join(problems + actions)

    if area in ["급여계약·인정서 상담", "방문요양 연계 상담", "병원입원·전원·퇴소 상담"] or any(word in selected_text for word in trigger_words):
        return (
            "후속 문서 작성 안내: 본 상담 내용은 욕구사정 및 급여제공계획 검토 대상에 해당할 수 있습니다. "
            "인정서 갱신, 등급변경, 급여계약 연장, 서비스 내용 변경, 방문요양 연계 또는 서비스 중단·퇴소 사유가 있는 경우 "
            "욕구사정 실시 여부와 급여제공계획서 갱신 필요 여부를 확인하시기 바랍니다."
        )

    return "후속 문서 작성 안내: 현재 상담 내용은 일반 상담기록으로 관리하되, 상태변화나 서비스 변경 필요성이 확인될 경우 욕구사정 및 급여제공계획 반영 여부를 검토하시기 바랍니다."


def make_sentence(area, problems, causes, actions, program_area, program_name, event_name):
    problem_text = ", ".join(problems)
    cause_text = ", ".join(causes)
    action_text = ", ".join(actions)

    extra = ""
    if area == "프로그램 참여 상담":
        extra = f" 프로그램 영역은 '{program_area}'이며, 프로그램명은 '{program_name}'으로 확인된다."
    elif area == "행사 참여 상담":
        extra = f" 행사명은 '{event_name}'으로 확인된다."

    consult = (
        f"{area}을 실시하였다.{extra}\n\n"
        f"상담 결과 주요 내용으로 {problem_text} 관련 사항이 확인되었다. "
        f"해당 내용은 수급자의 주간보호 이용 과정에서 관찰되거나 보호자 상담을 통해 확인된 사항으로, "
        f"서비스 이용 만족도와 일상생활 지원 방향을 점검하기 위해 상담을 진행하였다.\n\n"
        f"상담하게 된 원인으로는 {cause_text} 등이 고려되었다. "
        f"기관에서는 해당 사항이 일시적인 변화인지, 반복적인 지원 필요사항인지 확인하기 위해 상담내용을 기록하고 관련 직원과 공유하기로 하였다."
    )

    action = (
        f"조치사항으로 {action_text}을/를 실시하거나 검토하기로 하였다. "
        f"수급자의 안전과 서비스 만족도를 높이기 위해 관련 내용을 담당 직원과 공유하고, "
        f"필요 시 보호자에게 상담 결과를 안내하기로 하였다.\n\n"
        f"향후 동일한 문제가 반복되거나 서비스 내용 변경이 필요한 경우, 욕구사정 및 급여제공계획 변경 여부를 검토하기로 하였다."
    )

    reflection = (
        "상담을 통해 확인된 수급자 및 보호자의 요구사항은 주간보호 급여제공 과정에 반영 여부를 검토한다. "
        "식사, 송영, 프로그램, 행사 참여, 병원동행, 개인활동지원, 안전관리 등 서비스 내용에 반영 가능한 사항은 관련 직원과 공유하여 적용한다. "
        "변경사항이 없는 경우 기존 급여제공계획을 유지하되, 상태변화 또는 보호자 요청이 지속될 경우 급여제공계획 변경을 검토한다."
    )

    monitoring = (
        "사후 모니터링은 담당 직원이 이용 중 관찰을 통해 진행하며, 필요 시 보호자와 추가 상담을 실시한다. "
        "상담 이후 수급자의 상태 변화, 서비스 만족도, 프로그램 참여도, 보호자 요청사항 반영 여부를 다음 상담 시 재확인하기로 한다."
    )

    guardian_notice = (
        "보호자 전달사항: 상담 결과 및 기관 조치사항을 보호자에게 안내하고, 추가 요청사항이 있는 경우 기관으로 연락하도록 안내하였다."
    )

    followup_notice = make_followup_notice(area, problems, actions)

    return consult, action, reflection, monitoring, guardian_notice, followup_notice


def make_excel(data, consult, action, reflection, monitoring, guardian_notice, followup_notice):
    wb = Workbook()
    ws = wb.active
    ws.title = "주간보호 상담일지"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    current_row = 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws.cell(current_row, 1).value = f"( {data['기관명']} ) 주간보호센터 상담일지"
    ws.cell(current_row, 1).font = Font(bold=True, size=22)
    ws.cell(current_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 36
    current_row += 1

    info_rows = [
        ("수급자 성명", data["수급자 성명"], "생년월일", data["생년월일"], "성별", data["성별"]),
        ("장기요양 등급", data["장기요양 등급"], "이용시작일", str(data["이용시작일"]), "이용요일", data["이용요일"]),
        ("이용시간", data["이용시간"], "본인부담률", data["본인부담률"], "상담일자", str(data["상담일자"])),
        ("상담시간", str(data["상담시간"]), "상담구분", data["상담구분"], "상담장소", data["상담장소"]),
        ("상담방법", data["상담방법"], "상담대상자", data["상담대상자"], "관계", data["관계"]),
        ("담당자", data["담당자"], "직책", data["직책"], "동석자", data["동석자"]),
        ("상담영역", data["상담영역"], "프로그램/행사", data["프로그램/행사"], "비고", data["비고"]),
    ]

    for row_data in info_rows:
        ws.cell(current_row, 1).value = row_data[0]
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2).value = row_data[1]

        ws.cell(current_row, 4).value = row_data[2]
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        ws.cell(current_row, 5).value = row_data[3]

        ws.cell(current_row, 7).value = row_data[4]
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=9)
        ws.cell(current_row, 8).value = row_data[5]

        for col in [1, 4, 7]:
            ws.cell(current_row, col).fill = header_fill
            ws.cell(current_row, col).font = Font(bold=True, size=11, name="맑은 고딕")
            ws.cell(current_row, col).alignment = Alignment(horizontal="center", vertical="center")

        for col in [2, 5, 8]:
            ws.cell(current_row, col).font = Font(size=11, name="맑은 고딕")

        current_row += 1

    sections = [
        ("상담 내용", consult, 7),
        ("조치 내용", action, 6),
        ("급여제공 반영 여부", reflection, 5),
        ("사후 조치 및 모니터링 계획", monitoring, 4),
        ("보호자 전달사항", guardian_notice, 3),
        ("후속 문서 작성 안내", followup_notice, 3),
    ]

    for title, text, height in sections:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(current_row, 1).value = title
        ws.cell(current_row, 1).fill = header_fill
        ws.cell(current_row, 1).font = Font(bold=True, size=12, name="맑은 고딕")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="center")
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + height - 1, end_column=9)
        ws.cell(current_row, 1).value = text
        ws.cell(current_row, 1).font = Font(size=11, name="맑은 고딕")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        current_row += height

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=9)

    ws.cell(current_row, 1).value = "상담자 서명"
    ws.cell(current_row, 4).value = f"{data['담당자']}  (서명)"
    ws.cell(current_row, 7).value = "대상자/보호자  (서명)"

    for col in [1, 7]:
        ws.cell(current_row, col).fill = header_fill
        ws.cell(current_row, col).font = Font(bold=True, size=11, name="맑은 고딕")
        ws.cell(current_row, col).alignment = Alignment(horizontal="center")

    ws.cell(current_row, 4).font = Font(size=11, name="맑은 고딕")
    ws.cell(current_row, 4).alignment = Alignment(horizontal="right")

    max_row = current_row

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "left",
                vertical=cell.alignment.vertical or "center",
                wrap_text=True
            )

    for r in range(1, max_row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 24

    # 기본 인쇄 설정
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_options.horizontalCentered = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


col1, col2, col3 = st.columns(3)

with col1:
    facility = st.text_input("기관명", "BeeCare 주간보호센터")
    name = st.text_input("수급자 성명", "홍길동")
    birth = st.text_input("생년월일", "1940-01-01")
    gender = st.selectbox("성별", ["남", "여"])
    grade = st.selectbox("장기요양 등급", ["1등급", "2등급", "3등급", "4등급", "5등급", "인지지원등급"])

with col2:
    start_date = st.date_input("이용시작일", date.today())
    use_days = st.text_input("이용요일", "월, 수, 금")
    use_time = st.text_input("이용시간", "09:00~17:00")
    copay = st.selectbox("본인부담률", ["0%", "6%", "9%", "12%", "15%", "20%"])
    consult_date = st.date_input("상담일자", date.today())

with col3:
    consult_time = st.time_input("상담시간", time(10, 0))
    consult_type = st.selectbox("상담 구분", ["정기 상담", "수시 상담", "보호자 요청 상담", "계약 관련 상담", "상태변화 상담"])
    place = st.selectbox("상담 장소", ["상담실", "프로그램실", "송영차량", "유선상담", "기타"])
    method = st.selectbox("상담 방법", ["대면 면담", "유선/전화 상담", "온라인/SNS", "기타"])
    target = st.text_input("상담 대상자 성명", "보호자")

col4, col5, col6 = st.columns(3)

with col4:
    relation = st.selectbox("상담자와의 관계", ["본인", "아들", "딸", "배우자", "며느리", "사위", "기타"])
    staff_position = st.selectbox("기관 담당자 직책", ["사회복지사", "간호사", "간호조무사", "요양보호사", "시설장"])
    staff_name = st.text_input("기관 담당자 성명", "사회복지사")

with col5:
    companion = st.text_input("동석자", "없음")
    area = st.selectbox("상담영역 선택", list(AREA_DATA.keys()))
    memo = st.text_input("비고", "특이사항 없음")

with col6:
    program_area = st.selectbox("프로그램 영역", ["해당 없음", "인지활동", "신체활동", "음악활동", "미술활동", "회상활동", "여가활동", "특별활동"])
    program_name = st.text_input("프로그램명", "실버체조")
    event_name = st.text_input("행사명", "생일잔치")

st.subheader("문제상황 / 원인 / 조치 선택")

problems = st.multiselect("문제상황 선택", AREA_DATA[area]["problems"], default=AREA_DATA[area]["problems"][:2])
causes = st.multiselect("상담하게 된 원인 선택", AREA_DATA[area]["causes"], default=AREA_DATA[area]["causes"][:2])
actions = st.multiselect("조치내용 선택", AREA_DATA[area]["actions"], default=AREA_DATA[area]["actions"][:2])

consult, action, reflection, monitoring, guardian_notice, followup_notice = make_sentence(
    area=area,
    problems=problems,
    causes=causes,
    actions=actions,
    program_area=program_area,
    program_name=program_name,
    event_name=event_name,
)

st.subheader("상담 내용")
consult = st.text_area("상담내용 수정 가능", consult, height=220)

st.subheader("조치 내용")
action = st.text_area("조치내용 수정 가능", action, height=180)

st.subheader("급여제공 반영 여부")
reflection = st.text_area("급여제공 반영 내용 수정 가능", reflection, height=130)

st.subheader("사후 조치 및 모니터링 계획")
monitoring = st.text_area("모니터링 계획 수정 가능", monitoring, height=120)

st.subheader("보호자 전달사항")
guardian_notice = st.text_area("보호자 전달사항 수정 가능", guardian_notice, height=100)

st.subheader("후속 문서 작성 안내")
followup_notice = st.text_area("후속문서 안내 수정 가능", followup_notice, height=100)

checked_items = []
st.subheader("필수 기재항목 체크리스트 - 앱 화면 확인용")

for item in REQUIRED_ITEMS:
    if st.checkbox(item, value=True, key=f"check_{item}"):
        checked_items.append(item)

missing_items = [item for item in REQUIRED_ITEMS if item not in checked_items]
score = int((len(checked_items) / len(REQUIRED_ITEMS)) * 100)

if score == 100:
    st.success(f"상담일지 완성도: {score}점 / 누락 항목 없음")
elif score >= 85:
    st.warning(f"상담일지 완성도: {score}점 / 일부 항목 확인 필요")
else:
    st.error(f"상담일지 완성도: {score}점 / 필수항목 보완 필요")

program_or_event = ""
if area == "프로그램 참여 상담":
    program_or_event = f"{program_area} / {program_name}"
elif area == "행사 참여 상담":
    program_or_event = event_name
else:
    program_or_event = "해당 없음"

data = {
    "기관명": facility,
    "수급자 성명": name,
    "생년월일": birth,
    "성별": gender,
    "장기요양 등급": grade,
    "이용시작일": start_date,
    "이용요일": use_days,
    "이용시간": use_time,
    "본인부담률": copay,
    "상담일자": consult_date,
    "상담시간": consult_time,
    "상담구분": consult_type,
    "상담장소": place,
    "상담방법": method,
    "상담대상자": target,
    "관계": relation,
    "담당자": staff_name,
    "직책": staff_position,
    "동석자": companion,
    "상담영역": area,
    "프로그램/행사": program_or_event,
    "비고": memo,
}

excel_file = make_excel(data, consult, action, reflection, monitoring, guardian_notice, followup_notice)

st.download_button(
    label="📥 주간보호센터 상담일지 엑셀 다운로드",
    data=excel_file,
    file_name=f"{name}_{area}_주간보호상담일지.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
