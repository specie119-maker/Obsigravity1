import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, time

st.set_page_config(page_title="BeeCare 급여제공지침 교육일지 1", layout="centered")

st.title("BeeCare 급여제공지침 교육일지 생성기 1")
st.caption("감염예방 · 개인정보보호 · 근골격계 · 노인인권 · 직원 인권침해 · 고충처리 지침")

facility_name = st.text_input("기관명")
education_date = st.date_input("교육일자", value=date.today())
start_time = st.time_input("시작시간", value=time(13, 30))
end_time = st.time_input("종료시간", value=time(14, 30))
place = st.text_input("장소", value="사무실")
instructor = st.text_input("강사", value="시설장")
manager = st.text_input("담당자")
director = st.text_input("시설장 결재자")

GUIDELINES = {
    "감염 예방 및 관리 지침": {
        "purpose": "장기요양기관 내 감염 발생을 예방하고, 종사자가 손위생·개인보호구 사용·환경관리·감염 의심 시 보고절차를 숙지하여 어르신과 종사자의 안전을 보호함.",
        "content": """1. 감염 예방의 이해
- 감염의 의미
- 장기요양기관에서 감염관리가 중요한 이유
- 어르신은 면역력이 약해 감염에 취약함
2. 감염 위험요인
- 손위생 미흡
- 기침, 발열, 설사 등 감염 의심 증상
- 기저질환 및 면역저하
- 공동생활 공간 사용
- 식사, 배변, 목욕 도움 과정에서의 접촉
3. 감염 예방 기본수칙
- 서비스 전후 손 씻기 또는 손소독
- 마스크, 장갑 등 개인보호구 사용
- 기침 예절 준수
- 식사도움, 배변도움 후 손위생 실시
- 오염물 처리 시 장갑 착용
- 자주 접촉하는 물품과 공간 소독
4. 감염 의심 시 대응절차
- 발열, 기침, 설사, 피부병변 등 증상 확인
- 담당자 및 시설장에게 즉시 보고
- 필요 시 보호자 안내
- 다른 어르신과 접촉 최소화
- 기관 지침에 따른 소독 및 격리조치 협조
5. 종사자 실천사항
- 출근 전 건강상태 확인
- 감염 의심 증상 발생 시 즉시 보고
- 손위생과 마스크 착용 생활화
- 식기, 수건, 개인물품 구분 사용
- 감염관리 기록 유지
6. 교육 후 유의사항
- 감염예방은 매일 반복 실천이 중요함
- 손위생은 가장 기본적인 예방수칙임
- 의심 증상 발견 시 지체 없이 보고해야 함""",
        "material": """[첨부 교육자료: 감염 예방 및 관리 지침]
1. 감염 예방의 필요성
장기요양기관은 고령의 어르신이 함께 생활하거나 이용하는 공간이므로 감염 발생 시 빠르게 전파될 수 있다. 특히 어르신은 면역력이 약하고 만성질환을 가진 경우가 많아 감염 예방관리가 매우 중요하다.
2. 손위생
손위생은 감염예방의 가장 기본이다. 서비스 제공 전후, 식사도움 전후, 배변도움 후, 기저귀 교환 후, 코를 풀거나 기침한 후, 오염물을 만진 후에는 손 씻기 또는 손소독을 실시한다.
3. 개인보호구 사용
마스크, 장갑, 앞치마 등은 상황에 맞게 사용한다. 장갑을 착용하더라도 장갑 사용 전후 손위생을 실시해야 하며, 오염된 장갑으로 다른 물건을 만지지 않도록 주의한다.
4. 환경관리
문손잡이, 침상 난간, 휠체어 손잡이, 식탁, 화장실 손잡이 등 자주 접촉하는 부위는 정기적으로 청결하게 관리한다. 오염물은 정해진 절차에 따라 처리하고, 환기와 청소를 주기적으로 실시한다.
5. 감염 의심 증상 대응
발열, 기침, 인후통, 설사, 구토, 피부병변 등 감염 의심 증상을 발견하면 즉시 담당자에게 보고한다. 다른 어르신과의 접촉을 줄이고 기관 지시에 따라 보호자 안내, 진료연계, 소독 등 필요한 조치를 실시한다.
6. 교육 후 확인사항
- 손위생이 필요한 상황을 설명할 수 있는가?
- 개인보호구 사용 후 손위생의 필요성을 이해했는가?
- 감염 의심 증상 발견 시 보고절차를 알고 있는가?
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
    },

    "개인정보 보호 지침": {
        "purpose": "어르신의 개인정보와 민감정보를 안전하게 보호하고, 종사자가 기록·사진·영상·상담내용 관리 시 개인정보 유출을 예방하도록 함.",
        "content": """1. 개인정보의 이해
- 개인정보의 정의
- 장기요양 현장에서 다루는 개인정보
- 민감정보와 건강정보의 중요성
2. 보호해야 할 정보
- 이름, 생년월일, 주소, 연락처
- 장기요양등급, 건강상태, 질병정보
- 가족관계, 상담내용
- 사진, 영상, 급여제공기록
- 보호자 연락처와 개인 사정
3. 개인정보 보호 실무
- 기록지 외부 반출 금지
- 동의 없는 사진·영상 촬영 금지
- 카카오톡, 문자 전송 시 주의
- 불필요한 개인정보 공유 금지
- 서류 보관장소 관리
- 퇴사 후에도 비밀유지 의무 준수
4. 개인정보 유출 예방
- 컴퓨터와 휴대전화 잠금 설정
- 출력물 방치 금지
- 다른 이용자나 외부인에게 정보 제공 금지
- 교육자료나 홍보물 사용 시 동의 확인
5. 유출 의심 시 대응
- 즉시 담당자와 시설장에게 보고
- 유출된 정보 종류와 경위 확인
- 추가 유출 방지
- 필요한 경우 보호자 안내 및 기관 조치
6. 교육 후 유의사항
- 출처 없는 정보 공유 금지
- 사진과 영상은 반드시 동의 확인
- 개인정보는 최소한으로 사용하고 안전하게 보관해야 함""",
        "material": """[첨부 교육자료: 개인정보 보호 지침]
1. 개인정보의 의미
개인정보란 특정 개인을 알아볼 수 있는 모든 정보를 말한다. 장기요양기관에서는 어르신의 이름, 주소, 연락처, 건강상태, 장기요양등급, 가족관계, 상담내용, 사진, 영상, 급여제공기록 등이 개인정보에 해당한다.
2. 민감정보 보호
건강정보, 질병, 투약, 장애, 장기요양등급 등은 민감한 정보로 더욱 주의가 필요하다. 종사자는 업무상 알게 된 내용을 가족이 아닌 외부인이나 다른 이용자에게 말하지 않아야 한다.
3. 사진 및 영상 관리
어르신의 사진이나 영상을 촬영하거나 홍보자료, 교육자료, 단체 대화방 등에 사용하는 경우에는 반드시 동의를 확인해야 한다. 동의 없이 촬영하거나 공유하는 행위는 개인정보 침해가 될 수 있다.
4. 기록지 관리
급여제공기록지, 상담기록, 건강상태 기록 등은 정해진 장소에 보관하고 외부 반출을 금지한다. 출력물을 책상 위에 방치하지 않으며, 폐기 시에는 개인정보가 보이지 않도록 안전하게 처리한다.
5. 유출 시 대응
개인정보 유출이 의심되면 즉시 담당자와 시설장에게 보고한다. 유출된 정보의 종류, 경위, 대상자를 확인하고 추가 유출을 막아야 한다.
6. 교육 후 확인사항
- 개인정보에 해당하는 정보를 설명할 수 있는가?
- 사진과 영상 사용 시 동의가 필요함을 이해했는가?
- 개인정보 유출 의심 시 보고절차를 알고 있는가?
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
    },

    "근골격계 질환 예방 지침": {
        "purpose": "종사자가 어르신 이동보조, 체위변경, 목욕도움 등 업무 중 발생할 수 있는 근골격계 부담요인을 이해하고 안전한 작업자세를 실천하도록 함.",
        "content": """1. 근골격계 질환의 이해
- 근골격계 질환의 정의
- 허리, 어깨, 손목, 무릎 통증의 원인
- 장기요양 업무와 근골격계 부담
2. 발생 위험요인
- 무리한 들어올리기
- 허리를 굽힌 자세
- 반복적인 손목 사용
- 장시간 서 있는 업무
- 좁은 공간에서의 이동보조
- 혼자서 무리하게 대상자를 이동시키는 경우
3. 예방을 위한 작업자세
- 허리만 굽히지 않고 무릎을 사용
- 어르신을 몸 가까이에 두고 보조
- 방향 전환 시 허리를 비틀지 않음
- 체위변경 시 보조도구 활용
- 무리한 경우 동료에게 도움 요청
4. 업무별 예방수칙
- 침상 이동보조
- 휠체어 이동보조
- 목욕도움
- 기저귀 교환
- 체위변경
- 물품 이동
5. 종사자 건강관리
- 작업 전후 스트레칭
- 통증 발생 시 조기 보고
- 적절한 휴식
- 반복 작업 시 자세 변경
- 보호대나 보조기구 활용
6. 교육 후 유의사항
- 통증을 참으며 무리하지 않음
- 혼자 들어올리는 작업을 피함
- 종사자의 건강도 안전한 급여제공의 기본임""",
        "material": """[첨부 교육자료: 근골격계 질환 예방 지침]
1. 근골격계 질환의 의미
근골격계 질환은 반복적인 작업, 무리한 힘, 부적절한 자세 등으로 인해 허리, 목, 어깨, 팔, 손목, 무릎 등에 통증이나 기능저하가 발생하는 질환을 말한다.
2. 장기요양 업무의 위험요인
장기요양 종사자는 어르신 이동보조, 체위변경, 목욕도움, 기저귀 교환, 휠체어 이동 등 신체 부담이 큰 업무를 수행한다. 이때 허리를 과도하게 굽히거나 어르신을 혼자 들어올리면 부상의 위험이 높아진다.
3. 올바른 작업자세
어르신을 이동시킬 때는 허리만 굽히지 않고 무릎을 굽혀 몸 전체를 사용한다. 어르신을 몸 가까이에 두고, 방향을 바꿀 때는 허리를 비틀지 말고 발의 위치를 함께 옮긴다. 필요 시 동료의 도움을 요청한다.
4. 스트레칭과 건강관리
작업 전후로 어깨, 허리, 손목 스트레칭을 실시한다. 통증이 지속되거나 저림, 힘 빠짐이 있을 경우 담당자에게 보고하고 필요한 조치를 받는다.
5. 교육 후 확인사항
- 이동보조 시 허리만 굽히면 안 되는 이유를 이해했는가?
- 무리한 작업 시 도움을 요청해야 함을 알고 있는가?
- 통증 발생 시 보고해야 함을 이해했는가?
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
    },

    "학대 예방 및 대응 지침": {
    "purpose": "아동·노인·장애인 등 학대 취약계층에 대한 학대 유형을 이해하고 위기 상황 발생 시 신속한 대응 및 피해자 보호 조치를 수행할 수 있도록 하며, 관련 법령에 따른 신고의무자의 역할과 책임을 숙지하도록 함.",

    "content": """1. 교육 목적
- 아동·노인·장애인 등 학대 취약계층에 대한 학대 유형 이해
- 위기 상황 발생 시 신속한 대응 및 피해자 보호 조치 확립
- 관련 법령에 따른 신고의무자의 역할과 책임 인식
2. 교육 내용
① 학대의 정의 및 다양한 유형
- 신체적 학대
- 정서적 학대
- 성적 학대
- 경제적 학대
- 방임 및 유기
- 실제 사례를 통한 유형별 학대 사례 분석
② 학대 예방 지침 및 초기 대응 절차
- 학대 의심 징후 확인 방법
- 피해자 발견 시 접근 방법
- 피해자 안전 확보 조치
- 학대 의심 상황 발생 시 기관 내 보고 절차
- 기관 대응 매뉴얼 숙지
③ 신고의무자의 역할 및 신고 방법
- 신고의무자의 법적 책임
- 노인학대 신고 절차
- 노인보호전문기관 연계
- 경찰 신고 절차
- 신고자 보호제도
- 신고자 비밀유지 의무
④ 학대 피해자 보호 및 사후조치
- 피해자 분리 보호
- 응급치료 지원
- 보호자 상담 및 연계
- 피해자 개인정보 보호
- 기관의 후속 관리 및 재발방지 대책 수립
3. 교육 자료 및 준비물
[✔] 2026년 학대 예방 및 신고의무자 교육 교재
[✔] 기관 자체 학대 예방 및 대응 지침 매뉴얼
[✔] 노인보호전문기관 신고 안내자료
[✔] 학대 사례 교육자료
4. 종사자 실천사항
- 학대 의심 상황 발견 시 즉시 보고
- 피해자 안전 확보 우선
- 신고의무자의 역할 충실 이행
- 신고자 및 피해자 정보 비밀 유지
- 학대 예방을 위한 존중 중심 돌봄 실천
5. 교육 후 유의사항
- 학대는 반드시 신고해야 하는 중대한 사안임
- 학대 의심만으로도 신고 가능함
- 신고자 보호와 비밀유지는 중요함
- 학대 예방은 모든 종사자의 공동 책임임""",

    "material": """[첨부 교육자료 : 2026년 학대 예방 및 대응 지침]
교육명
2026년 학대 예방 및 대응 지침 교육
교육대상
본 기관 종사자 및 신고의무자 전원
교육장소
기관 내 사무실 또는 온라인 화상회의 플랫폼
교육방법
[✔] 집합 대면 교육
[ ] 시청각 교육(온라인)
[ ] 화상 교육
1. 학대의 정의
학대란 보호가 필요한 아동·노인·장애인 등 취약계층에게 신체적, 정서적, 성적, 경제적 피해를 주거나 필요한 보호와 돌봄을 제공하지 않는 행위를 말한다.
2. 학대의 유형
① 신체적 학대
폭행, 밀침, 강압적인 신체 제지 등
② 정서적 학대
폭언, 무시, 협박, 모욕, 따돌림 등
③ 성적 학대
동의 없는 신체접촉, 성적 수치심 유발 행위
④ 경제적 학대
금전 갈취, 재산 부당 사용 등
⑤ 방임 및 유기
식사, 위생, 의료 등 필요한 돌봄을 제공하지 않는 행위
3. 학대 의심 징후
- 설명하기 어려운 멍, 상처
- 위축, 불안, 두려움
- 영양불량과 위생불량
- 통장 및 금전 관련 이상 호소
- 보호자 또는 특정인을 두려워함
4. 발견 시 대응절차
- 어르신 안전 확보
- 피해자 비난 금지
- 사실 중심 관찰 및 기록
- 담당자 및 시설장 보고
- 필요 시 관계기관 신고
5. 종사자 실천사항
- 어르신의 말을 경청
- 학대 의심상황 은폐 금지
- 개인정보 보호
- 재발방지 관찰
6. 교육 후 유의사항
- 의심만 되어도 보고 필요
- 판단보다 보호가 우선
- 기록은 객관적으로 작성
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
},
    "직원 인권 침해 대응 지침": {
        "purpose": "종사자가 업무 중 폭언, 폭행, 성희롱, 부당한 요구 등 인권침해 상황에 노출될 경우 안전하게 대응하고 보호받을 수 있도록 함.",
        "content": """1. 직원 인권의 이해
- 종사자도 보호받아야 할 권리가 있음
- 안전한 근무환경의 중요성
- 감정노동과 인권침해 예방
2. 직원 인권침해 유형
- 폭언과 욕설
- 폭행 또는 위협
- 성희롱 및 성적 불쾌감 유발
- 과도한 사적 요구
- 부당한 업무지시
- 인격무시와 괴롭힘
3. 발생 상황별 대응
- 즉시 안전거리 확보
- 혼자 대응하지 않음
- 담당자 및 시설장 보고
- 필요 시 동료 동행
- 심각한 경우 경찰 또는 관계기관 연계
4. 기관의 보호조치
- 피해 종사자 상담
- 업무조정
- 재발방지 안내
- 보호자 또는 이용자와의 조정
- 기록관리
5. 종사자 실천사항
- 위협 상황을 참거나 숨기지 않음
- 사실 중심으로 기록
- 반복 발생 시 즉시 보고
- 동료의 피해도 함께 보호
6. 교육 후 유의사항
- 종사자의 안전이 확보되어야 좋은 돌봄이 가능함
- 인권침해는 개인 문제가 아니라 기관이 관리해야 할 사항임""",
        "material": """[첨부 교육자료: 직원 인권 침해 대응 지침]
1. 직원 인권의 의미
장기요양 종사자는 어르신에게 돌봄서비스를 제공하는 전문인력이며, 업무 중에도 안전하고 존중받을 권리가 있다. 폭언, 폭행, 성희롱, 부당한 요구는 종사자의 인권을 침해하는 행위이다.
2. 인권침해 유형
종사자에게 욕설이나 모욕적인 말을 하는 경우, 신체적 위협을 가하는 경우, 성적 농담이나 접촉을 하는 경우, 업무 범위를 벗어난 사적 심부름을 반복적으로 요구하는 경우 등이 해당될 수 있다.
3. 대응원칙
위협을 느끼면 즉시 안전거리를 확보하고 혼자 대응하지 않는다. 담당자나 시설장에게 즉시 보고하며, 필요 시 동료와 함께 대응한다. 심각한 폭행이나 위협이 있는 경우 관계기관의 도움을 받을 수 있다.
4. 보고 및 기록
발생일시, 장소, 상대방, 발언 또는 행동, 목격자, 조치내용을 사실 중심으로 기록한다. 반복되는 상황은 기관 차원의 보호조치와 재발방지 대책이 필요하다.
5. 교육 후 확인사항
- 직원 인권침해 유형을 설명할 수 있는가?
- 위협 상황에서 혼자 대응하지 않아야 함을 이해했는가?
- 피해 발생 시 보고와 기록이 필요함을 알고 있는가?
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
    },

    "고충 처리 지침": {
        "purpose": "어르신, 보호자, 종사자의 고충을 체계적으로 접수·처리하여 서비스 질을 개선하고 권익을 보호하도록 함.",
        "content": """1. 고충처리의 이해
- 고충의 의미
- 고충처리의 필요성
- 이용자와 종사자의 권익보호
2. 고충의 유형
- 서비스 불만
- 종사자 응대 관련 불만
- 식사, 위생, 환경 관련 불편
- 보호자 요구사항
- 직원 근무 관련 고충
- 인권침해 관련 고충
3. 고충 접수 방법
- 구두 접수
- 서면 접수
- 전화 또는 보호자 상담
- 고충처리함
- 정기 상담 및 회의
4. 처리절차
- 고충 접수
- 내용 확인
- 담당자 검토
- 처리방안 마련
- 결과 안내
- 재발방지 조치
- 기록 보관
5. 종사자 실천사항
- 불만을 가볍게 넘기지 않음
- 방어적으로 대응하지 않음
- 경청하고 기록함
- 담당자에게 보고
- 개인정보 보호
6. 교육 후 유의사항
- 고충은 서비스 개선의 기회임
- 처리과정은 객관적으로 기록
- 결과 안내와 재발방지가 중요함""",
        "material": """[첨부 교육자료: 고충 처리 지침]
1. 고충처리의 의미
고충처리는 어르신, 보호자, 종사자가 서비스 이용 또는 근무 과정에서 느끼는 불편, 불만, 건의사항을 접수하고 해결하는 절차를 말한다. 고충처리는 기관의 서비스 질 개선과 권익보호를 위해 필요하다.
2. 고충의 종류
어르신의 식사, 위생, 프로그램, 환경, 종사자 응대에 대한 불만, 보호자의 요구사항, 직원의 업무 부담이나 인권침해 관련 호소 등이 고충에 포함될 수 있다.
3. 접수방법
고충은 구두, 전화, 상담, 서면, 고충처리함, 회의 등을 통해 접수할 수 있다. 종사자는 어르신이나 보호자의 불편을 들었을 때 가볍게 넘기지 않고 담당자에게 보고해야 한다.
4. 처리절차
고충이 접수되면 내용을 확인하고 담당자가 검토한다. 필요한 경우 관련자 면담을 실시하고 처리방안을 마련한다. 처리결과는 가능한 범위에서 신청자에게 안내하며, 재발방지 대책을 마련한다.
5. 기록관리
고충 접수일, 신청자, 내용, 처리과정, 결과, 재발방지 조치를 기록한다. 개인정보와 민감한 내용은 외부에 유출되지 않도록 보호한다.
6. 교육 후 확인사항
- 고충 접수방법을 설명할 수 있는가?
- 고충을 담당자에게 보고해야 함을 이해했는가?
- 처리결과와 재발방지 기록의 중요성을 알고 있는가?
※ 본 교육자료는 장기요양기관 급여제공지침 및 관련 평가기준을 참고하여 BeeCare에서 교육용으로 재구성한 자료입니다."""
    },
}

RESULTS = {
    "감염 예방 및 관리 지침": "손위생, 개인보호구 사용, 환경관리 및 감염 의심 시 보고절차를 숙지하도록 교육을 실시함.",
    "개인정보 보호 지침": "개인정보와 민감정보 보호, 사진·영상 관리, 기록물 보관 및 유출 예방수칙을 숙지하도록 교육을 실시함.",
    "근골격계 질환 예방 지침": "안전한 이동보조, 올바른 작업자세, 스트레칭 및 통증 발생 시 보고절차를 숙지하도록 교육을 실시함.",
    "학대 예방 및 대응 지침": "학대 유형별 특성과 대응절차, 신고의무자의 역할 및 신고방법을 숙지하도록 교육을 실시함.",
    "직원 인권 침해 대응 지침": "폭언, 폭행, 성희롱, 부당요구 등 직원 인권침해 상황의 대응 및 보고절차를 숙지하도록 교육을 실시함.",
    "고충 처리 지침": "고충 접수, 처리절차, 결과 안내, 재발방지 및 기록관리 방법을 숙지하도록 교육을 실시함.",
}

st.subheader("지침 선택")
if "selected_guidelines" not in st.session_state:
    st.session_state.selected_guidelines = []

cols = st.columns(2)
for idx, item in enumerate(GUIDELINES.keys()):
    selected_now = item in st.session_state.selected_guidelines
    label = f"✅ {item}" if selected_now else item
    if cols[idx % 2].button(label, key=f"guide_{idx}", use_container_width=True):
        if selected_now:
            st.session_state.selected_guidelines.remove(item)
        else:
            st.session_state.selected_guidelines.append(item)
        st.rerun()

selected = st.session_state.selected_guidelines

if selected:
    st.success(f"{len(selected)}개 지침이 선택되었습니다.")

st.subheader("참석자 명단")
roles = ["시설장", "사회복지사", "간호사/간호조무사", "요양보호사", "물리치료사", "기타"]

attendees = []
for role in roles:
    names = st.text_input(f"{role} 이름", placeholder="예: 김OO, 이OO")
    if names:
        attendees.append((role, names))


def style_sheet(ws, border):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical=cell.alignment.vertical or "center",
                wrap_text=True
            )
            if cell.value and cell.font.bold:
                continue
            if cell.row != 1:
                cell.font = Font(size=11)


def create_excel():
    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "교육일지"

    for col in range(1, 10):  # A~I만 사용
        ws.column_dimensions[get_column_letter(col)].width = 16

      # 제목 + 결재라인
    ws.merge_cells("A1:G2")
    ws["A1"] = "급여제공지침 교육일지"
    ws["A1"].font = Font(size=30, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["H1"] = "담당자"
    ws["I1"] = "시설장"
    ws["H2"] = manager
    ws["I2"] = director

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 55

    rows = [
        ("교육명", " / ".join(selected) + " 교육"),
        ("교육일시", f"{education_date} {start_time} ~ {end_time}"),
        ("장소", place),
        ("강사", instructor),
    ]

    r = 3
    for title, value in rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(r, 1).value = title
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        ws.cell(r, 3).value = value
        ws.row_dimensions[r].height = 25
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1).value = "교육목적"
    ws.cell(r, 1).font = Font(size=20, bold=True)
    ws.row_dimensions[r].height = 30
    r += 1

    purpose_text = "\n\n".join([f"[{item}]\n{GUIDELINES[item]['purpose']}" for item in selected])
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=9)
    ws.cell(r, 1).value = purpose_text
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 5

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1).value = "주요 교육내용"
    ws.cell(r, 1).font = Font(size=20, bold=True)
    ws.row_dimensions[r].height = 30
    r += 1

    content_text = "\n\n".join([f"[{item}]\n{GUIDELINES[item]['content']}" for item in selected])
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 16, end_column=9)
    ws.cell(r, 1).value = content_text
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 17

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1).value = "교육결과"
    ws.cell(r, 1).font = Font(size=20, bold=True)
    ws.row_dimensions[r].height = 30
    r += 1

    result_text = "\n".join([f"- {RESULTS[item]}" for item in selected])
    result_text = (
        "급여제공지침에 대해 종사자 교육을 실시하였으며,\n"
        + result_text +
        "\n\n교육 후 종사자는 지침별 주요 예방수칙과 대응절차를 확인하였으며, "
        "실제 급여제공 과정에서 적용할 수 있도록 교육을 실시함."
    )

    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=9)
    ws.cell(r, 1).value = result_text
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 5

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1).value = "참석자 명단"
    ws.cell(r, 1).font = Font(bold=True)
    ws.row_dimensions[r].height = 25
    r += 1

    ws.cell(r, 1).value = "직책"
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r, 2).value = "성명"
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
    ws.cell(r, 6).value = "서명"
    r += 1

    for role, names in attendees:
        ws.cell(r, 1).value = role
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(r, 2).value = names
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        ws.cell(r, 6).value = ""
        ws.row_dimensions[r].height = 24
        r += 1

    # 첨부자료 시트
    ws2 = wb.create_sheet("첨부자료")
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "급여제공지침 첨부 교육자료"
    ws2["A1"].font = Font(size=22, bold=True)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    row = 3
    for item in selected:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws2.cell(row, 1).value = item
        ws2.cell(row, 1).font = Font(size=15, bold=True)
        ws2.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[row].height = 28
        row += 1

        text = (
            f"교육목적\n{GUIDELINES[item]['purpose']}\n\n"
            f"주요 교육내용\n{GUIDELINES[item]['content']}\n\n"
            f"{GUIDELINES[item]['material']}"
        )
        ws2.merge_cells(start_row=row, start_column=1, end_row=row + 22, end_column=8)
        ws2.cell(row, 1).value = text
        ws2.cell(row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 24

    # 교육결과 시트
    ws3 = wb.create_sheet("교육결과")
    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    ws3.merge_cells("A1:H1")
    ws3["A1"] = "교육결과"
    ws3["A1"].font = Font(size=22, bold=True)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40

    ws3.merge_cells("A3:H8")
    ws3["A3"] = result_text
    ws3["A3"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws3.merge_cells("A10:H14")
    ws3["A10"] = (
        "평가 대비 권장사항\n"
        "- 교육사진 1~2장 첨부\n"
        "- 참석자 자필서명 확인\n"
        "- 교육자료 함께 편철\n"
        "- 교육일자, 강사, 참석자, 교육내용 누락 여부 확인"
    )
    ws3["A10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    style_sheet(ws, border)
    style_sheet(ws2, border)
    style_sheet(ws3, border)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


if st.button("교육일지 생성"):
    if not selected:
        st.warning("급여제공지침을 1개 이상 선택해주세요.")
    else:
        excel_file = create_excel()
        st.success("교육일지가 생성되었습니다.")
        st.download_button(
            label="엑셀 다운로드",
            data=excel_file,
            file_name="beecare_pay_edu2_교육일지.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
