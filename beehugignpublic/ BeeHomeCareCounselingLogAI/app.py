from datetime import date
import streamlit as st

# 하루 3회 제한 설정
DAILY_LIMIT = 3

if "usage_date" not in st.session_state:
    st.session_state.usage_date = date.today()

if "download_count" not in st.session_state:
    st.session_state.download_count = 0

# 날짜가 바뀌면 횟수 초기화
if st.session_state.usage_date != date.today():
    st.session_state.usage_date = date.today()
    st.session_state.download_count = 0

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date

st.set_page_config(page_title="방문요양센터 상담일지 생성기", layout="centered")

st.title("방문요양센터 상담일지 생성기")
st.caption("보호자 상담, 요양보호사 상담, 시간변경, 퇴소, 지역자원연계까지 방문요양 전용으로 작성합니다.")

# -----------------------------
# 상담 유형 템플릿
# -----------------------------
consult_templates = {
    "정기상담": {
        "content": "수급자의 전반적인 건강상태, 서비스 만족도, 생활환경, 제공시간 적절성 및 보호자 의견을 확인함.",
        "action": "현재 제공 중인 방문요양서비스를 유지하며, 특이사항 발생 시 보호자 및 담당자에게 즉시 공유하기로 함.",
        "need_plan": False,
    },

    "보호자 상담 - 초기 상담 및 욕구사정": {
        "content": "서비스 시작 전 보호자와 초기 상담을 실시함. 수급자의 신체적·인지적 건강상태, 질환 정보, 생활환경, 보호자가 원하는 돌봄 방향 및 특이사항을 확인함. 방문요양 급여 범위, 본인부담금, 이용 절차를 안내함.",
        "action": "상담 내용을 바탕으로 초기 욕구사정 및 급여제공계획 수립에 반영하기로 함.",
        "need_plan": True,
    },

    "보호자 상담 - 정기 및 수시 상담": {
        "content": "서비스 제공 중 보호자와 정기 또는 수시 상담을 실시함. 요양보호사 서비스 만족도, 수급자의 건강상태 변화, 낙상·건강악화·갈등 등 특이사항, 센터 운영일정 및 장기요양보험 안내사항을 확인함.",
        "action": "보호자 의견과 수급자 상태 변화를 기록하고, 필요한 경우 요양보호사와 공유하여 서비스 제공내용을 조정하기로 함.",
        "need_plan": False,
    },

    "보호자 상담 - 결과평가 및 급여계획 변경": {
        "content": "보호자 의견을 수렴하여 지난 서비스 제공 결과를 평가함. 수급자에게 제공된 방문요양서비스가 적절했는지, 욕구 변화나 보호자 요청사항이 있는지 확인함.",
        "action": "상담 결과를 바탕으로 급여제공계획 변경 필요 여부를 검토하고, 수급자의 욕구 변화 또는 보호자 요청사항이 확인될 경우 30일 이내 급여제공계획을 재작성하기로 함.",
        "need_plan": True,
    },

    "요양보호사 상담 - 직무 및 근무환경 관리": {
        "content": "요양보호사의 업무 범위, 방문 일정, 서비스 제공 현황에 대해 상담함. 가족을 위한 가사노동 등 급여 범위를 벗어난 요청 여부와 수급자 상태 변화에 따른 서비스 조정 필요성을 확인함.",
        "action": "업무 범위를 재안내하고, 근무일정 또는 서비스 제공계획 변경이 필요한 경우 사회복지사가 검토하기로 함.",
        "need_plan": False,
    },

    "요양보호사 상담 - 고충 및 스트레스 상담": {
        "content": "요양보호사가 서비스 제공 중 경험한 고충사항과 스트레스 요인에 대해 상담함. 수급자 및 보호자와의 관계, 부당한 요구, 감정소진 여부를 확인함.",
        "action": "애로사항을 청취하고 필요 시 기관 차원의 지원, 보호자 상담 또는 근무 조정을 검토하기로 함.",
        "need_plan": False,
    },

    "요양보호사 상담 - 건강 및 안전·감염 관리": {
        "content": "근골격계 질환 예방, 감염관리, 낙상예방 및 안전사고 대응에 대해 상담함. 서비스 제공 중 발생 가능한 위험요인을 확인함.",
        "action": "안전수칙 및 감염예방수칙을 재안내하고, 건강상태와 근무 중 위험요인을 지속 관찰하기로 함.",
        "need_plan": False,
    },

    "요양보호사 상담 - 직업윤리 및 역량 강화": {
        "content": "노인학대 예방, 비밀유지 의무, 개인정보 보호, 수급자 존중 및 직업윤리에 대해 상담함. 치매 어르신 응대법 등 직무교육 필요 여부를 확인함.",
        "action": "관련 교육자료를 제공하고 직무교육 참여를 안내하기로 함.",
        "need_plan": False,
    },

    "건강상태 변화": {
        "content": "수급자의 건강상태 변화가 확인되어 상담함. 식사, 수면, 통증, 보행, 인지상태, 복약 여부 등을 확인함.",
        "action": "건강상태 변화를 지속 관찰하고 필요 시 보호자에게 병원진료를 안내하며 급여제공계획 변경 여부를 검토하기로 함.",
        "need_plan": True,
    },

    "서비스 시간변경 요청": {
        "content": "수급자 또는 보호자가 방문요양 서비스 제공시간 변경을 요청하여 상담함. 기존 제공시간, 변경 희망시간, 요일 추가 여부 및 변경 사유를 확인함.",
        "action": "월 한도액, 본인부담금, 요양보호사 근무 가능 여부를 확인한 후 급여제공계획 변경 필요 여부를 검토하기로 함.",
        "need_plan": True,
    },

    "요양보호사 변경요청": {
        "content": "수급자 또는 보호자가 요양보호사 변경을 요청하여 상담함. 변경요청 사유와 현재 서비스 제공 중 불편사항을 확인함.",
        "action": "변경 사유를 확인하고 요양보호사 배정 가능 여부를 검토한 후 보호자에게 안내하기로 함.",
        "need_plan": False,
    },

    "병원진료": {
        "content": "수급자의 병원진료 일정과 진료 필요사항에 대해 상담함.",
        "action": "진료 결과에 따라 수급자 상태 변화를 확인하고 필요 시 서비스 제공내용 조정 여부를 검토하기로 함.",
        "need_plan": False,
    },

    "한의원진료": {
        "content": "수급자의 한의원진료 이용과 관련하여 보호자 및 수급자 의견을 확인함.",
        "action": "진료 후 상태 변화를 관찰하고 필요 시 보호자와 추가 상담하기로 함.",
        "need_plan": False,
    },

    "병원입원": {
        "content": "수급자의 병원입원으로 인해 방문요양서비스 이용 조정이 필요함을 상담함.",
        "action": "입원기간 동안 급여제공 중단 또는 일정 조정 여부를 확인하고, 퇴원 후 재상담을 실시하기로 함.",
        "need_plan": True,
    },

    "퇴원 후 재가복귀 상담": {
        "content": "수급자의 병원 퇴원 후 재가복귀와 방문요양서비스 재개 필요사항에 대해 상담함.",
        "action": "퇴원 후 건강상태와 돌봄 필요도를 확인하고 급여제공계획 변경 필요 여부를 검토하기로 함.",
        "need_plan": True,
    },

    "퇴소상담": {
        "content": "수급자의 서비스 이용 종료 사유를 확인하고 퇴소 관련 상담을 실시함. 퇴소 사유, 종료일, 보호자 의견 및 향후 서비스 이용계획을 확인함.",
        "action": "퇴소 사유에 따라 급여제공 종료일을 확인하고 관련 기록 및 서류를 정리하기로 함.",
        "need_plan": True,
    },

    "욕구사정 실시": {
        "content": "연 2회 이상 실시하는 욕구사정과 관련하여 수급자 및 보호자 의견을 확인함.",
        "action": "욕구사정 결과를 바탕으로 급여제공계획 반영 여부를 검토하기로 함.",
        "need_plan": True,
    },

    "급여제공계획 변경": {
        "content": "수급자의 상태 변화 또는 보호자 요청에 따라 급여제공계획 변경 필요성에 대해 상담함.",
        "action": "상담 내용을 바탕으로 욕구사정 및 급여제공계획서 갱신 필요 여부를 확인하기로 함.",
        "need_plan": True,
    },

    "인정서 재발급": {
        "content": "장기요양인정서 재발급과 관련하여 상담함.",
        "action": "재발급된 인정서 내용을 확인한 후 급여제공계획서 변경 필요 여부를 검토하기로 함.",
        "need_plan": True,
    },

    "등급변경": {
        "content": "수급자의 장기요양등급 변경과 관련하여 상담함.",
        "action": "변경된 등급에 따라 욕구사정 및 급여제공계획서를 갱신하기로 함.",
        "need_plan": True,
    },

    "5등급 인지활동 워크북 확인": {
        "content": "5등급 수급자의 인지활동형 방문요양 제공과 관련하여 워크북 제공 여부, 활용 정도, 참여도 및 수행 가능 수준을 확인함.",
        "action": "워크북 활용 내용을 지속 확인하고 요양보호사에게 인지활동 제공기록을 충실히 작성하도록 안내하기로 함.",
        "need_plan": False,
    },

    "지역자원 연계": {
        "content": "수급자의 생활상 어려움과 지역사회 자원 연계 필요 여부를 상담함. 기초생활수급, 노인복지관, 행정복지센터, 보건소 등 이용 가능 자원을 확인함.",
        "action": "필요한 지역자원을 안내하고 신청 또는 이용 여부를 확인하여 센터 상담기록에 반영하기로 함.",
        "need_plan": False,
    },

    "도시락 연계": {
        "content": "수급자의 식사 준비 어려움과 영양상태를 확인하고 도시락 지원 필요 여부를 상담함.",
        "action": "도시락 지원 가능 기관 또는 지역자원을 확인하여 보호자에게 안내하고 연계 여부를 확인하기로 함.",
        "need_plan": False,
    },

    "예방접종 안내 및 확인": {
        "content": "수급자의 예방접종 일정, 접종 필요성 및 접종 여부를 확인함.",
        "action": "보건소 또는 의료기관 예방접종 정보를 안내하고 접종 여부를 추후 확인하기로 함.",
        "need_plan": False,
    },
}

# -----------------------------
# 세부 사유 목록
# -----------------------------
time_change_reasons = [
    "기존 3시간 제공에서 4시간 제공으로 변경 요청",
    "주 5회에서 주 6회로 요일 추가 요청",
    "보호자 부재시간 증가로 서비스 시간 확대 필요",
    "식사준비·청소·세탁 등 가사활동지원 시간이 부족함",
    "목욕, 이동도움, 병원동행 등 신체활동지원 시간이 추가로 필요함",
    "수급자 상태 변화로 돌봄 시간이 더 필요함",
]

caregiver_change_reasons = [
    "수급자와 요양보호사 간 성향 및 의사소통이 맞지 않음",
    "서비스 제공시간 조정이 어려움",
    "수급자 또는 보호자의 서비스 만족도 저하",
    "요양보호사의 개인사정 또는 근무 지속 어려움",
    "신체활동지원·가사활동지원 등 서비스 제공방식에 대한 의견 차이",
]

discharge_reasons = [
    "타 기관 전원",
    "병원 입원",
    "요양병원 입소",
    "시설 입소",
    "수급자 사망",
    "보호자 요청",
    "장기 부재",
    "기타",
]

workbook_items = [
    "워크북을 제공받았는지 확인함",
    "워크북 활동 참여도와 흥미도를 확인함",
    "인지활동 자료 활용 내용과 반복활동 필요성을 안내함",
]

resource_items = [
    "기초생활수급 신청을 안내하고 등록 여부를 확인함",
    "노인복지관 취미활동 프로그램 연계를 안내함",
    "예방접종 일정을 안내하고 접종 여부를 확인함",
    "도시락 지원 연계 가능성을 확인함",
    "행정복지센터, 보건소, 복지관 등 지역자원 이용 가능 여부를 확인함",
]

# -----------------------------
# 입력 화면
# -----------------------------
center_name = st.text_input("기관명", value="")
client_name = st.text_input("수급자명", value="")
birth = st.text_input("생년월일", value="")
consult_date = st.date_input("상담일자", value=date.today())
consult_time = st.text_input("상담시간", value="10:00")
writer = st.text_input("작성자", value="사회복지사")

consult_method = st.selectbox(
    "상담방법",
    ["방문상담", "전화상담", "보호자 상담", "요양보호사 상담", "기타"]
)

consult_target = st.selectbox(
    "상담대상",
    ["수급자", "보호자", "수급자+보호자", "요양보호사", "보호자+요양보호사", "기타"]
)

selected_type = st.selectbox("상담유형", list(consult_templates.keys()))

detail_text = ""

if selected_type == "서비스 시간변경 요청":
    selected_details = st.multiselect("시간변경 요청 세부사유", time_change_reasons)
    detail_text = " / ".join(selected_details)

elif selected_type == "요양보호사 변경요청":
    selected_details = st.multiselect("요양보호사 변경요청 사유", caregiver_change_reasons)
    detail_text = " / ".join(selected_details)

elif selected_type == "퇴소상담":
    selected_detail = st.selectbox("퇴소 사유", discharge_reasons)
    detail_text = selected_detail

elif selected_type == "5등급 인지활동 워크북 확인":
    selected_details = st.multiselect("워크북 확인 내용", workbook_items)
    detail_text = " / ".join(selected_details)

elif selected_type in ["지역자원 연계", "도시락 연계", "예방접종 안내 및 확인"]:
    selected_details = st.multiselect("지역자원 연계 내용", resource_items)
    detail_text = " / ".join(selected_details)

default_content = consult_templates[selected_type]["content"]
default_action = consult_templates[selected_type]["action"]

if detail_text:
    default_content += f"\n\n세부내용: {detail_text}"

consult_content = st.text_area("상담내용", value=default_content, height=170)
action_content = st.text_area("조치사항", value=default_action, height=120)
extra_note = st.text_area("특이사항 / 추가기록", value="", height=80)

need_plan = consult_templates[selected_type]["need_plan"]

if need_plan:
    st.warning("이 상담유형은 욕구사정 또는 급여제공계획서 갱신 여부 확인이 필요합니다.")

include_plan_notice = st.checkbox(
    "하단에 욕구사정 및 급여제공계획서 갱신 필요 안내 삽입",
    value=need_plan
)

# -----------------------------
# 엑셀 생성
# -----------------------------
def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "상담일지"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="EDEDED")

    title_font = Font(name="맑은 고딕", size=22, bold=True)
    header_font = Font(name="맑은 고딕", size=11, bold=True)
    normal_font = Font(name="맑은 고딕", size=11)

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    current_row = 1

    # 결재라인 H/I 분리
    ws.cell(current_row, 8, "담당자")
    ws.cell(current_row, 9, "시설장")
    ws.cell(current_row + 1, 8, "")
    ws.cell(current_row + 1, 9, "")

    for r in range(current_row, current_row + 2):
        for c in range(8, 10):
            cell = ws.cell(r, c)
            cell.border = border
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row += 3

    # 제목
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    ws.cell(current_row, 1, f"({center_name}) 방문요양 상담일지")
    ws.cell(current_row, 1).font = title_font
    ws.cell(current_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 36
    current_row += 1

    # 기본정보
    info_rows = [
        ("수급자명", client_name, "생년월일", birth, "상담일자", str(consult_date)),
        ("상담시간", consult_time, "상담방법", consult_method, "상담대상", consult_target),
        ("상담유형", selected_type, "작성자", writer, "기관명", center_name),
    ]

    for row in info_rows:
        ws.cell(current_row, 1, row[0])
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        ws.cell(current_row, 2, row[1])

        ws.cell(current_row, 4, row[2])
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        ws.cell(current_row, 5, row[3])

        ws.cell(current_row, 7, row[4])
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=9)
        ws.cell(current_row, 8, row[5])

        for c in range(1, 10):
            cell = ws.cell(current_row, c)
            cell.border = border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c in [1, 4, 7]:
                cell.fill = fill
                cell.font = header_font

        ws.row_dimensions[current_row].height = 28
        current_row += 1

    sections = [
        ("상담내용", consult_content),
        ("조치사항", action_content),
        ("특이사항 / 추가기록", extra_note if extra_note else "해당 없음"),
    ]

    if include_plan_notice:
        sections.append((
            "욕구사정 및 급여제공계획 연계 안내",
            "본 상담내용에 따라 욕구사정 및 급여제공계획서 갱신 필요 여부를 확인해야 함."
        ))

    for title, content in sections:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(current_row, 1, title)
        ws.cell(current_row, 1).fill = fill
        ws.cell(current_row, 1).font = header_font
        ws.cell(current_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, 10):
            ws.cell(current_row, c).border = border

        ws.row_dimensions[current_row].height = 26
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(current_row, 1, content)
        ws.cell(current_row, 1).font = normal_font
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for c in range(1, 10):
            ws.cell(current_row, c).border = border

        ws.row_dimensions[current_row].height = 90
        current_row += 1

    # 작성자 확인
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws.cell(current_row, 1, "위와 같이 상담을 실시하고 기록함.")

    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=9)
    ws.cell(current_row, 7, f"작성자: {writer}  (서명)")

    for c in range(1, 10):
        cell = ws.cell(current_row, c)
        cell.border = border
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[current_row].height = 32

    # 인쇄설정
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_area = f"A1:I{current_row}"

    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return file

# -----------------------------
# 다운로드
# -----------------------------
if st.button("상담일지 엑셀 생성"):
    excel_file = create_excel()
    filename = f"{client_name if client_name else '수급자'}_방문요양_상담일지.xlsx"

    st.download_button(
        label="엑셀 다운로드",
        data=excel_file,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
